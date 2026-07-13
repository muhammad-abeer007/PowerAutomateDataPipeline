"""
DHU Hourly Report ETL Pipeline
==============================
Processes "Factory_Fin_DHU_Hourly_MM_dd_yyyy_HH_mm_ss.xlsx" files, transforming
each one (unmerge/forward-fill, drop Total / Defect Ranking / Grand Total rows,
drop blank columns, unpivot defect categories, extract filename metadata,
derive Date/Time/Hourname columns), and writes one transformed file per input
file to an output folder.

The input path can be EITHER:
  - A folder: every matching file inside it is processed.
  - A single .xlsx file: only that file is processed.

Run from CMD:
    python dhu_etl.py
    python dhu_etl.py "C:/path/to/input_folder" "C:/path/to/output"
    python dhu_etl.py "C:/path/to/single_file.xlsx" "C:/path/to/output"

If no arguments are given, the script uses the INPUT_PATH / OUTPUT_FOLDER
variables below, prompting for the input path at runtime if left blank. If no
output folder is supplied, an "Output" subfolder is created inside the input
folder (or inside the single file's parent folder).
"""

import os
import re
import sys
import datetime
import pandas as pd
import openpyxl

# ----------------------------------------------------------------------
# EDITABLE DEFAULTS -- leave blank ("") to be prompted for the path at runtime
# INPUT_PATH may be a folder OR a single .xlsx file
# ----------------------------------------------------------------------
INPUT_PATH = r""
OUTPUT_FOLDER = r""

# ----------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------
FILENAME_PATTERN = re.compile(
    r"^(?P<factory>[A-Za-z0-9]+)_Fin_DHU_Hourly_"
    r"(?P<mm>\d{2})_(?P<dd>\d{2})_(?P<yyyy>\d{4})_"
    r"(?P<HH>\d{2})_(?P<mm2>\d{2})_(?P<ss>\d{2})$",
    re.IGNORECASE,
)

SHEET_NAME = "Sheet1"

DROP_KEYWORDS = ("total", "defect ranking", "grand total")

# Canonical fixed-column names, keyed by a normalized (whitespace-collapsed,
# lowercased) version of the header text so that minor spacing differences
# between source files are tolerated but the OUTPUT header text always
# matches the ETL demo exactly.
FIXED_COLUMNS_CANONICAL = {
    "prod group": "Prod Group",
    "time slot": "Time Slot",
    "defect qty": "Defect Qty",
    "defective piece qty": "Defective Piece Qty",
    "pass qty (output)": "Pass Qty (Output)",
    "output qty": "Output Qty",
    "prod qty (1st check)": "Prod  Qty (1st Check)",
    "checked qty": "Checked Qty",
    "reject qty": "Reject Qty",
    "fpy (%)": "FPY (%)",
    "dhu(%)": "DHU(%)",
    "dhu (%)": "DHU(%)",
    "reject (%)": "Reject (%)",
}

FINAL_COLUMN_ORDER = [
    "Factory", "Date_Time", "Date", "Time",
    "Prod Group", "Time Slot",
    "Defect Qty", "Defective Piece Qty", "Pass Qty (Output)", "Output Qty",
    "Prod  Qty (1st Check)", "Checked Qty", "Reject Qty",
    "FPY (%)", "DHU(%)", "Reject (%)",
    "Attribute", "Value", "Time-Hourname", "Hourname",
]


def normalize(text):
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip().lower()


def find_header_row(ws):
    """Locate the row whose cells contain 'Prod Group' (the real header row)."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if normalize(cell.value) == "prod group":
                return cell.row
    return None


def build_header_map(ws, header_row):
    """Map column index -> header text for every non-blank header cell."""
    header_map = {}
    for cell in ws[header_row]:
        if cell.value is not None and str(cell.value).strip() != "":
            header_map[cell.column] = str(cell.value).strip()
    return header_map


def get_hour_24(hour_token, ampm):
    """Convert a Time Slot's leading hour token (+ optional AM/PM) to 24h."""
    h = int(hour_token)
    if h <= 12:
        if ampm and ampm.upper() == "PM" and h != 12:
            return h + 12
        if ampm and ampm.upper() == "AM" and h == 12:
            return 0
        return h
    return h  # already > 12, i.e. effectively 24-hour style


def hourname_from_time_slot(time_slot):
    """Derive the 1-13 Hourname code from a Time Slot string."""
    if not time_slot or not isinstance(time_slot, str):
        return None
    m = re.match(r"\s*(\d{1,2}):\d{2}\s*([AP]M)?", time_slot, re.IGNORECASE)
    if not m:
        return None
    hour24 = get_hour_24(m.group(1), m.group(2))
    code = hour24 - 7
    return code if 1 <= code <= 13 else None


def hourname_from_time(time_val):
    """Derive the Time-Hourname code (1-13) from a datetime.time value."""
    if time_val is None:
        return None
    code = time_val.hour - 7
    return code if 1 <= code <= 13 else None


def extract_filename_metadata(stem):
    m = FILENAME_PATTERN.match(stem)
    if not m:
        return None
    factory = m.group("factory")
    date_time = f"{m.group('mm')}_{m.group('dd')}_{m.group('yyyy')}_{m.group('HH')}_{m.group('mm2')}_{m.group('ss')}"
    date_val = datetime.date(int(m.group("yyyy")), int(m.group("mm")), int(m.group("dd")))
    time_val = datetime.time(int(m.group("HH")), int(m.group("mm2")), int(m.group("ss")))
    return factory, date_time, date_val, time_val


def transform_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"'{SHEET_NAME}' not found (sheets present: {wb.sheetnames})")
    ws = wb[SHEET_NAME]

    header_row = find_header_row(ws)
    if header_row is None:
        raise ValueError("Could not locate the 'Prod Group' header row")

    header_map = build_header_map(ws, header_row)

    records = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        rec = {}
        for cell in row:
            if cell.column in header_map:
                rec[header_map[cell.column]] = cell.value
        if rec:
            records.append(rec)

    df = pd.DataFrame(records)
    if "Prod Group" not in df.columns:
        raise ValueError("Prod Group column missing after parsing header")

    # Unmerge / forward-fill the vertically merged Prod Group cells
    df["Prod Group"] = df["Prod Group"].ffill()

    # Drop rows with no Time Slot (covers Total / Defect Ranking / Grand Total
    # rows, and any stray blank rows)
    df = df[df["Time Slot"].notna()]

    # Defensive: also explicitly drop any row whose Prod Group / Time Slot
    # text contains the disallowed keywords
    def is_dropped(val):
        n = normalize(val)
        return any(k in n for k in DROP_KEYWORDS)

    df = df[~df["Prod Group"].apply(is_dropped)]
    df = df[~df["Time Slot"].apply(is_dropped)]

    # Split columns into fixed (canonical) vs defect-category (attribute) columns
    rename_map = {}
    fixed_present = set()
    for col in df.columns:
        key = normalize(col)
        if key in FIXED_COLUMNS_CANONICAL:
            rename_map[col] = FIXED_COLUMNS_CANONICAL[key]
            fixed_present.add(FIXED_COLUMNS_CANONICAL[key])
    df = df.rename(columns=rename_map)

    attribute_cols = [c for c in df.columns if c not in fixed_present]

    id_vars = [c for c in ["Prod Group", "Time Slot", "Defect Qty", "Defective Piece Qty",
                            "Pass Qty (Output)", "Output Qty", "Prod  Qty (1st Check)",
                            "Checked Qty", "Reject Qty", "FPY (%)", "DHU(%)", "Reject (%)"]
               if c in df.columns]

    df = df.reset_index(drop=True)
    df["_row_id"] = df.index

    melted = df.melt(id_vars=id_vars + ["_row_id"], value_vars=attribute_cols,
                      var_name="Attribute", value_name="Value")
    melted = melted.dropna(subset=["Value"])

    # Rows where every defect-category column was blank produce no melted
    # output above -- but the row still had real production/check quantities
    # (e.g. Prod  Qty (1st Check)) that must feed the overall DHU% total, so
    # add an explicit "No Defects" / 0 record for each such row.
    rows_with_defects = set(melted["_row_id"])
    no_defect_row_ids = [rid for rid in df["_row_id"] if rid not in rows_with_defects]
    if no_defect_row_ids:
        no_defect_rows = df[df["_row_id"].isin(no_defect_row_ids)][id_vars + ["_row_id"]].copy()
        no_defect_rows["Attribute"] = "No Defects"
        no_defect_rows["Value"] = 0
        melted = pd.concat([melted, no_defect_rows], ignore_index=True)

    melted["Attribute"] = pd.Categorical(
        melted["Attribute"], categories=attribute_cols + ["No Defects"], ordered=True
    )
    melted = melted.sort_values(["_row_id", "Attribute"]).drop(columns=["_row_id"])
    melted["Attribute"] = melted["Attribute"].astype(str)

    # Filename metadata
    stem = os.path.splitext(os.path.basename(filepath))[0]
    meta = extract_filename_metadata(stem)
    if meta is None:
        raise ValueError(f"Filename '{stem}' does not match expected pattern")
    factory, date_time, date_val, time_val = meta

    melted["Factory"] = factory
    melted["Date_Time"] = date_time
    melted["Date"] = date_val
    melted["Time"] = time_val
    melted["Time-Hourname"] = hourname_from_time(time_val)
    melted["Hourname"] = melted["Time Slot"].apply(hourname_from_time_slot)

    for col in FINAL_COLUMN_ORDER:
        if col not in melted.columns:
            melted[col] = pd.NA

    return melted[FINAL_COLUMN_ORDER].reset_index(drop=True)


def resolve_paths():
    # Input path (file or folder): CLI arg > editable variable > interactive prompt (required)
    if len(sys.argv) >= 2 and sys.argv[1].strip():
        input_path = sys.argv[1].strip()
    else:
        input_path = INPUT_PATH
    if not input_path:
        input_path = input("Enter the input folder or file path: ").strip().strip('"')
    input_path = input_path.strip().strip('"')

    # Output folder: CLI arg > editable variable > auto-created "Output" subfolder
    # (never prompts here -- an unset output path is a valid, expected case)
    if len(sys.argv) >= 3 and sys.argv[2].strip():
        output_folder = sys.argv[2].strip()
    else:
        output_folder = OUTPUT_FOLDER.strip() if OUTPUT_FOLDER else ""
    output_folder = output_folder.strip().strip('"') if output_folder else ""

    if not output_folder:
        # Default "Output" folder lives beside the input: inside the folder
        # itself, or inside the parent folder of a single input file.
        base_dir = input_path if os.path.isdir(input_path) else os.path.dirname(input_path)
        output_folder = os.path.join(base_dir, "Output")

    return input_path, output_folder


def gather_matched_files(folder):
    matched = []
    for fname in os.listdir(folder):
        if fname.startswith("~$") or not fname.lower().endswith(".xlsx"):
            continue
        stem = os.path.splitext(fname)[0]
        if FILENAME_PATTERN.match(stem):
            matched.append(fname)
    return matched


def process_files(file_paths, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    ok, failed = 0, 0
    for src in sorted(file_paths):
        fname = os.path.basename(src)
        dst = os.path.join(output_folder, fname)
        try:
            result_df = transform_file(src)
            result_df.to_excel(dst, index=False, sheet_name="Output")
            print(f"  [OK]   {fname} -> {dst} ({len(result_df)} rows)")
            ok += 1
        except Exception as exc:
            print(f"  [FAIL] {fname}: {exc}")
            failed += 1
    print(f"\nDone. {ok} succeeded, {failed} failed. Output folder: {output_folder}")


def main():
    input_path, output_folder = resolve_paths()

    if os.path.isdir(input_path):
        matched_files = gather_matched_files(input_path)
        if not matched_files:
            print(f"No files matching 'Factory_Fin_DHU_Hourly_MM_dd_yyyy_HH_mm_ss.xlsx' found in {input_path}")
            sys.exit(0)
        print(f"Found {len(matched_files)} matching file(s) in folder. Processing...")
        file_paths = [os.path.join(input_path, f) for f in matched_files]
        process_files(file_paths, output_folder)

    elif os.path.isfile(input_path):
        if not input_path.lower().endswith(".xlsx"):
            print(f"ERROR: Not an .xlsx file: {input_path}")
            sys.exit(1)
        fname = os.path.basename(input_path)
        stem = os.path.splitext(fname)[0]
        if not FILENAME_PATTERN.match(stem):
            print(f"WARNING: '{fname}' does not match the expected "
                  f"'Factory_Fin_DHU_Hourly_MM_dd_yyyy_HH_mm_ss.xlsx' naming pattern; "
                  f"attempting to process it anyway.")
        print(f"Single file input detected. Processing '{fname}'...")
        process_files([input_path], output_folder)

    else:
        print(f"ERROR: Input path not found (not a valid file or folder): {input_path}")
        sys.exit(1)


if __name__ == "__main__":
    main()
