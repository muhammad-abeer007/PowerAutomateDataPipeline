r"""
Append the latest PGCL DHU hourly extract into the "Data_Source" Excel Table
inside the master workbook.

- Reads all data rows from the latest extract file (first/only sheet).
- Finds the Table named "Data_Source" in the master workbook (whichever
  sheet it lives on).
- For any Date already inside that table that also appears in the new file,
  those old rows are deleted (so re-running the same day's file refreshes
  it instead of duplicating rows).
- Appends the new file's rows to the end of the table and resizes the
  table's range to include them.
- Saves back to the master file path, overwriting it in place.

Run from cmd (single line, both paths quoted):
    python append_to_master.py "path\to\latest.xlsx" "path\to\master.xlsm"
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

TABLE_NAME = "Data_Source"
DATE_COL_NAME = "Date"


def get_path(prompt, arg_value):
    path = arg_value if arg_value else input(prompt).strip().strip('"')
    if not os.path.isfile(path):
        print(f"ERROR: file not found: {path}")
        sys.exit(1)
    return path


def read_new_rows(latest_path):
    wb = load_workbook(latest_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [
        list(r)
        for r in ws.iter_rows(min_row=2, values_only=True)
        if any(v is not None for v in r)
    ]
    return header, rows


def find_table(wb, table_name):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if table_name in ws.tables:
            return ws, ws.tables[table_name]
    return None, None


def main():
    latest_path = get_path("Path to latest hourly file: ", sys.argv[1] if len(sys.argv) > 1 else None)
    master_path = get_path("Path to master file: ", sys.argv[2] if len(sys.argv) > 2 else None)

    new_header, new_rows = read_new_rows(latest_path)
    if not new_rows:
        print("No data rows found in the latest file. Nothing to do.")
        return

    is_xlsm = master_path.lower().endswith(".xlsm")
    wb = load_workbook(master_path, data_only=False, keep_vba=is_xlsm)

    ws, table = find_table(wb, TABLE_NAME)
    if table is None:
        print(f"ERROR: Table '{TABLE_NAME}' not found in any sheet of the master file.")
        sys.exit(1)

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    master_header = [ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]

    if master_header != new_header:
        print("WARNING: column headers differ between the two files.")
        print(f"  Master table: {master_header}")
        print(f"  Latest file : {new_header}")

    col_index = {name: i for i, name in enumerate(new_header)}
    try:
        reordered_rows = [
            [row[col_index[name]] for name in master_header] for row in new_rows
        ]
    except KeyError as e:
        print(f"ERROR: column {e} present in master table but missing in latest file. Aborting.")
        sys.exit(1)

    date_col_pos = master_header.index(DATE_COL_NAME)  # 0-based within table columns
    new_dates = {row[date_col_pos] for row in reordered_rows}

    date_col_num = min_col + date_col_pos
    rows_to_delete = [
        r for r in range(min_row + 1, max_row + 1)
        if ws.cell(row=r, column=date_col_num).value in new_dates
    ]
    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

    data_end_row = max_row - len(rows_to_delete)
    for i, row in enumerate(reordered_rows):
        for j, value in enumerate(row):
            ws.cell(row=data_end_row + 1 + i, column=min_col + j, value=value)

    new_max_row = data_end_row + len(reordered_rows)
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
    table.ref = new_ref
    if table.autoFilter is not None:
        table.autoFilter.ref = new_ref

    wb.save(master_path)

    print(f"Removed {len(rows_to_delete)} old row(s) matching date(s): {sorted(str(d) for d in new_dates)}")
    print(f"Appended {len(reordered_rows)} new row(s).")
    print(f"Table '{TABLE_NAME}' now spans {new_ref}.")
    print(f"Master file updated and saved: {master_path}")


if __name__ == "__main__":
    main()
